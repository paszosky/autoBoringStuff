import openpyxl

workbook = openpyxl.load_workbook('arkusz.xlsx')
print(type(workbook))

sheet = workbook.get_sheet_by_name('Arkusz1')
print(workbook.get_sheet_names())

cell = sheet['A1']
print(cell.value)

print(sheet['C2'].value)
print(sheet.cell(row=2, column=3).value) #inny sposob wyboru

for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)